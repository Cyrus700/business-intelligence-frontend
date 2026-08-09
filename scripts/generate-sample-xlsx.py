"""Generate ready-to-upload sample XLSX workbooks for the data-integration panel.

Verify with:   python scripts/generate-sample-xlsx.py
Each workbook's first sheet matches the backend domain contract
(app.services.etl.domains.DOMAIN_SPECS) so uploads validate cleanly.
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parents[1] / "public" / "samples"

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
README_FONT = Font(size=10)


def _style_header(ws, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _add_readme(wb: Workbook, title: str, columns: list[str], notes: list[str]) -> None:
    ws = wb.create_sheet("README")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    cell = ws.cell(row=3, column=1, value="Required columns (exact names):")
    cell.font = Font(bold=True)
    row = 4
    for col in columns:
        ws.cell(row=row, column=1, value=f"• {col}").font = README_FONT
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True)
    row += 1
    for note in notes:
        ws.cell(row=row, column=1, value=f"• {note}").font = README_FONT
        row += 1
    ws.column_dimensions["A"].width = 60


def _save(wb: Workbook, name: str) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / name
    wb.save(path)
    print(f"wrote {path} ({path.stat().st_size} bytes)")


def sales() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sales"
    headers = [
        "date", "sku", "product_name", "category", "customer",
        "segment", "city", "region", "channel", "quantity", "unit_price", "discount",
    ]
    rows = [
        (date(2026, 7, 1), "SBX-1001", "Basmati Rice 25kg", "Staples", "Janaki Store", "retail", "Janakpur", "Madhesh", "store", 4, 3600, 0),
        (date(2026, 7, 1), "SBX-1002", "Marigold Garland (10pk)", "Festival", "Himalayan e-Shop", "online", "Pokhara", "Gandaki", "online", 12, 300, 15),
        (date(2026, 7, 2), "SBX-1003", "Wai Wai Noodles (30pk)", "Snacks", "Bhaktapur Super Store", "wholesale", "Bhaktapur", "Bagmati", "store", 200, 640, 0),
        (date(2026, 7, 3), "SBX-1004", "Detergent Powder 3kg", "Household", "Sundar Kirana", "retail", "Butwal", "Lumbini", "store", 6, 620, 31),
        (date(2026, 7, 4), "SBX-1005", "5 Star Chocolate (48pk)", "Confectionery", "Mountain Bazar", "online", "Kathmandu", "Bagmati", "online", 24, 1500, 75),
        (date(2026, 7, 5), "SBX-1006", "Nilkamal Chair (Set of 4)", "Furniture", "Chitwan Traders", "wholesale", "Bharatpur", "Bagmati", "store", 30, 5600, 0),
    ]
    _style_header(ws, headers)
    for row in rows:
        ws.append(row)
    _add_readme(
        wb,
        "Sales sample — upload with domain: Sales",
        ["date", "sku", "quantity", "unit_price"],
        [
            "quantity must be a whole number ≥ 1",
            "unit_price ≥ 0; discount ≤ line total",
            "date between 2000-01-01 and today",
        ],
    )
    _save(wb, "sales-sample.xlsx")


def finance() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "expenses"
    headers = ["date", "category", "amount", "department", "description"]
    rows = [
        (date(2026, 7, 1), "rent", 185000, "operations", "Warehouse & office rent"),
        (date(2026, 7, 2), "logistics", 13348.18, "operations", "Delivery & fuel"),
        (date(2026, 7, 3), "salaries", 380000, "admin", "Payroll — 14 employees"),
        (date(2026, 7, 5), "marketing", 25000, "sales", "Online campaign (July)"),
        (date(2026, 7, 8), "utilities", 4210.55, "operations", "Electricity & water"),
        (date(2026, 7, 10), "other", 750, "admin", "Stationery"),
    ]
    _style_header(ws, headers)
    for row in rows:
        ws.append(row)
    _add_readme(
        wb,
        "Finance — upload this stage: Finance",
        ["date", "category", "amount"],
        [
            "category must be one of: rent, salaries, utilities, marketing, logistics, other",
            "amount > 0",
        ],
    )
    _save(wb, "finance-sample.xlsx")


def inventory() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "inventory"
    headers = ["date", "sku", "quantity_on_hand", "reorder_level", "warehouse"]
    rows = [
        (date(2026, 7, 31), "SBX-1001", 522, 120, "main"),
        (date(2026, 7, 31), "SBX-1002", 554, 120, "main"),
        (date(2026, 7, 31), "SBX-1003", 1208, 300, "east"),
        (date(2026, 7, 31), "SBX-1004", 96, 200, "east"),
        (date(2026, 7, 31), "SBX-1005", 415, 100, "main"),
    ]
    _style_header(ws, headers)
    for row in rows:
        ws.append(row)
    _add_readme(
        wb,
        "Inventory — upload this stage: Inventory",
        ["date", "sku", "quantity_on_hand"],
        ["quantity_on_hand and reorder_level are whole numbers ≥ 0"],
    )
    _save(wb, "inventory-sample.xlsx")


def main() -> None:
    sales()
    finance()
    inventory()


if __name__ == "__main__":
    sys.exit(main())